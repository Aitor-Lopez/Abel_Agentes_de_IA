"""
evidence_reader.py
──────────────────
Lee todos los archivos de la carpeta de evidencias con Python puro
y devuelve su contenido como texto listo para inyectar en los prompts.
Soporta: .txt, .md, .json, .csv, .log, .pdf, .docx, .xlsx, .html
"""

import os
import json
from pathlib import Path

MAX_CHARS_PER_FILE = 12_000   # Límite por archivo para no saturar el contexto
MAX_TOTAL_CHARS    = 50_000   # Límite total de todas las evidencias combinadas


def _read_txt(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _read_pdf(path: Path) -> str:
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        return "\n".join(text_parts)
    except Exception:
        pass
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        return "\n".join(p.extract_text() or "" for p in reader.pages)
    except Exception as e:
        return f"[Error leyendo PDF: {e}]"


def _read_docx(path: Path) -> str:
    try:
        import docx
        doc = docx.Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        return f"[Error leyendo DOCX: {e}]"


def _read_xlsx(path: Path) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        rows = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows.append(f"## Hoja: {sheet}")
            for row in ws.iter_rows(values_only=True):
                if any(c is not None for c in row):
                    rows.append("\t".join(str(c) if c is not None else "" for c in row))
        return "\n".join(rows)
    except Exception as e:
        return f"[Error leyendo XLSX: {e}]"


def _read_json(path: Path) -> str:
    try:
        data = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        return json.dumps(data, ensure_ascii=False, indent=2)
    except Exception:
        return path.read_text(encoding="utf-8", errors="replace")


READERS = {
    ".txt":  _read_txt,
    ".md":   _read_txt,
    ".log":  _read_txt,
    ".csv":  _read_txt,
    ".html": _read_txt,
    ".json": _read_json,
    ".pdf":  _read_pdf,
    ".docx": _read_docx,
    ".doc":  _read_docx,
    ".xlsx": _read_xlsx,
    ".xls":  _read_xlsx,
}


def leer_evidencias(evidencias_dir: str | Path) -> dict[str, str]:
    """
    Lee todos los archivos soportados del directorio y devuelve
    un dict { nombre_archivo: contenido_texto }.
    """
    evidencias_dir = Path(evidencias_dir)
    resultado = {}

    for archivo in sorted(evidencias_dir.iterdir()):
        if not archivo.is_file():
            continue
        ext = archivo.suffix.lower()
        reader = READERS.get(ext)
        if reader is None:
            resultado[archivo.name] = f"[Formato no soportado: {ext}]"
            continue
        try:
            texto = reader(archivo)
            # Truncar si es demasiado largo
            if len(texto) > MAX_CHARS_PER_FILE:
                texto = texto[:MAX_CHARS_PER_FILE] + f"\n\n[... TRUNCADO - {len(texto)} chars totales ...]"
            resultado[archivo.name] = texto
        except Exception as e:
            resultado[archivo.name] = f"[Error al leer {archivo.name}: {e}]"

    return resultado


def formatear_evidencias_para_prompt(evidencias: dict[str, str]) -> str:
    """
    Convierte el dict de evidencias en un bloque de texto
    estructurado y listo para insertar en el prompt de un agente.
    """
    if not evidencias:
        return "⚠️ NO HAY ARCHIVOS EN LA CARPETA DE EVIDENCIAS."

    bloques = []
    total_chars = 0

    for nombre, contenido in evidencias.items():
        bloque = f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ARCHIVO: {nombre}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{contenido}
"""
        total_chars += len(bloque)
        if total_chars > MAX_TOTAL_CHARS:
            bloques.append(f"\n[... LÍMITE ALCANZADO - resto de archivos omitidos para no saturar el contexto ...]")
            break
        bloques.append(bloque)

    return "\n".join(bloques)
